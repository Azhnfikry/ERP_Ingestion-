"""
GHG Scope Emission Classifier v2 — Aethera Steel (AETH)
Processes SAP PO / DO / FI + commuting survey.
Adds Supplier Spend Analysis tab ranked by spend with cumulative %.
"""
import os, warnings
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint

warnings.filterwarnings("ignore")

# ── CONFIG ──────────────────────────────────────────────────────────────────────
REPORTING_YEAR  = 2025
WORKING_WEEKS   = 48
USD_TO_MYR      = 4.47
MYR_PER_KWH     = 0.509
MYR_PER_DIESEL  = 2.30
MYR_PER_FOIL    = 2.80
MYR_PER_MMBTU   = 21.0
OUTPUT          = "/mnt/user-data/outputs/ghg_classified_output_v2.xlsx"

EF = dict(
    diesel_litre=2.686, fuel_oil_litre=3.179, nat_gas_mmbtu=56.1,
    met_coke_t=3960.0, electricity_kwh=0.585,
    iron_ore_t=25.0, scrap_t=1.0, quicklime_t=785.0, dolomite_t=477.0,
    ferro_si_t=4000.0, ferro_mn_t=1800.0, liq_oxy_m3=0.26, nitrogen_m3=0.10,
    transport_tkm=0.096, car_km=0.170, motorcycle_km=0.103, bus_pax_km=0.089,
    steel_t=1830.0, water_litre=0.000344,
)

ROUTES = dict(KUL=50,PNG=320,JHR=360,PRK=200,SGR=40,SBH=1500,SWK=800,SLG=90)

def sf(v):
    try: return float(v)
    except: return 0.0

def find(name):
    for d in [".", "/mnt/user-data/uploads"]:
        p = os.path.join(d, name)
        if os.path.exists(p): return p
    raise FileNotFoundError(name)

# ── PROCESSORS ──────────────────────────────────────────────────────────────────

def proc_fi(path):
    df = pd.read_excel(path)
    out = []
    for _, r in df.iterrows():
        gl  = str(r.get("GL_ACCOUNT_TEXT",""))
        amt = sf(r.get("WRBTR",0))
        rec = dict(
            Source="SAP FI",
            Doc_Ref=str(r.get("BELNR","")),
            Date=str(r.get("BUDAT","")),
            Month=int(sf(r.get("MONAT",0)) or 0),
            Cost_Centre=str(r.get("KOSTL","")),
            Plant=str(r.get("WERKS","")),
            Vendor=str(r.get("NAME1","")),
            Description=gl,
            Spend_MYR=round(amt,2),
            Currency="MYR",
        )
        if "Electricity" in gl:
            kwh=amt/MYR_PER_KWH; kg=kwh*EF["electricity_kwh"]
            rec.update(Scope="Scope 2", GHG_Category="Purchased Electricity — TNB Grid",
                       Qty=round(kwh,2), Unit="kWh (est.)", EF_val=EF["electricity_kwh"],
                       kgCO2e=round(kg,2), tCO2e=round(kg/1000,4))
        elif "Natural Gas" in gl:
            mb=amt/MYR_PER_MMBTU; kg=mb*EF["nat_gas_mmbtu"]
            rec.update(Scope="Scope 1", GHG_Category="Stationary Combustion — Natural Gas",
                       Qty=round(mb,2), Unit="MMBTU (est.)", EF_val=EF["nat_gas_mmbtu"],
                       kgCO2e=round(kg,2), tCO2e=round(kg/1000,4))
        elif "Diesel" in gl:
            lt=amt/MYR_PER_DIESEL; kg=lt*EF["diesel_litre"]
            rec.update(Scope="Scope 1", GHG_Category="Mobile/Stationary Combustion — Diesel",
                       Qty=round(lt,2), Unit="Litres (est.)", EF_val=EF["diesel_litre"],
                       kgCO2e=round(kg,2), tCO2e=round(kg/1000,4))
        elif "Fuel Oil" in gl:
            lt=amt/MYR_PER_FOIL; kg=lt*EF["fuel_oil_litre"]
            rec.update(Scope="Scope 1", GHG_Category="Stationary Combustion — Fuel Oil",
                       Qty=round(lt,2), Unit="Litres (est.)", EF_val=EF["fuel_oil_litre"],
                       kgCO2e=round(kg,2), tCO2e=round(kg/1000,4))
        elif "Water" in gl:
            lt=amt/0.003; kg=lt*EF["water_litre"]
            rec.update(Scope="Scope 3 Cat 5", GHG_Category="Waste — Water & Wastewater Treatment (Air Selangor)",
                       Qty=round(lt,0), Unit="Litres (est.)", EF_val=EF["water_litre"],
                       kgCO2e=round(kg,2), tCO2e=round(kg/1000,4))
        else:
            rec.update(Scope="Unclassified", GHG_Category="Review Required",
                       Qty=0, Unit="", EF_val=0, kgCO2e=0, tCO2e=0)
        out.append(rec)
    return pd.DataFrame(out)


def proc_po(path):
    df = pd.read_excel(path)
    out = []
    for _, r in df.iterrows():
        mg    = str(r.get("Material Group","")).strip()
        short = str(r.get("Short Text","")).strip().upper()
        qty   = sf(r.get("PO Quantity",0))
        unit  = str(r.get("OUn","")).strip()
        val   = sf(r.get("Net Order Value",0))
        curr  = str(r.get("Currency","MYR")).strip()
        vmyr  = val * USD_TO_MYR if curr=="USD" else val
        rec   = dict(
            Source="SAP PO",
            Doc_Ref=f"{r.get('Purch.Doc.','')} / {r.get('Item','')}",
            Date=str(r.get("Doc. Date","")),
            Month="",
            Cost_Centre="",
            Plant=str(r.get("Plant","")),
            Vendor=str(r.get("Vendor Name","")),
            Description=str(r.get("Short Text","")),
            Material=str(r.get("Material","")),
            Material_Group=mg,
            Qty=qty,
            Unit=unit,
            Spend_MYR=round(vmyr,2),
            Currency=curr,
            Country=str(r.get("Country","")),
        )
        if mg == "ENERGY":
            if "DIESEL" in short:
                kg=qty*EF["diesel_litre"]
                rec.update(Scope="Scope 1", GHG_Category="Direct Combustion — Diesel (PO)",
                           EF_val=EF["diesel_litre"], kgCO2e=round(kg,2), tCO2e=round(kg/1000,4))
            elif "NGAS" in short or "NATURAL GAS" in short:
                kg=qty*EF["nat_gas_mmbtu"]
                rec.update(Scope="Scope 1", GHG_Category="Direct Combustion — Natural Gas (PO)",
                           EF_val=EF["nat_gas_mmbtu"], kgCO2e=round(kg,2), tCO2e=round(kg/1000,4))
            else:
                rec.update(Scope="Scope 1", GHG_Category="Energy Purchase — Review",
                           EF_val=0, kgCO2e=0, tCO2e=0)
        elif mg == "FUEL-RED":
            kg=qty*EF["met_coke_t"]
            rec.update(Scope="Scope 1", GHG_Category="Direct Combustion — Metallurgical Coke (EAF)",
                       EF_val=EF["met_coke_t"], kgCO2e=round(kg,2), tCO2e=round(kg/1000,4))
        elif mg == "RAW-IRON":
            kg=qty*EF["iron_ore_t"]
            rec.update(Scope="Scope 3 Cat 1", GHG_Category="Purchased Goods — Iron Ore",
                       EF_val=EF["iron_ore_t"], kgCO2e=round(kg,2), tCO2e=round(kg/1000,4))
        elif mg == "SCRAP":
            kg=qty*EF["scrap_t"]
            rec.update(Scope="Scope 3 Cat 1", GHG_Category="Purchased Goods — Steel Scrap",
                       EF_val=EF["scrap_t"], kgCO2e=round(kg,2), tCO2e=round(kg/1000,4))
        elif mg == "FLUX":
            ef_k = "quicklime_t" if "LIME" in short else "dolomite_t"
            kg=qty*EF[ef_k]
            rec.update(Scope="Scope 3 Cat 1", GHG_Category="Purchased Goods — Flux Materials",
                       EF_val=EF[ef_k], kgCO2e=round(kg,2), tCO2e=round(kg/1000,4))
        elif mg == "ALLOY":
            ef_k = "ferro_si_t" if ("FESI" in short or "SILICON" in short) else "ferro_mn_t"
            kg=qty*EF[ef_k]
            rec.update(Scope="Scope 3 Cat 1", GHG_Category="Purchased Goods — Ferroalloys",
                       EF_val=EF[ef_k], kgCO2e=round(kg,2), tCO2e=round(kg/1000,4))
        elif mg == "IND-GAS":
            ef_k = "liq_oxy_m3" if "OXY" in short else "nitrogen_m3"
            kg=qty*EF[ef_k]
            rec.update(Scope="Scope 3 Cat 1", GHG_Category="Purchased Goods — Industrial Gases",
                       EF_val=EF[ef_k], kgCO2e=round(kg,2), tCO2e=round(kg/1000,4))
        elif mg == "PACK":
            rec.update(Scope="Scope 3 Cat 1", GHG_Category="Purchased Goods — Packaging (no EF)",
                       EF_val="N/A", kgCO2e=0, tCO2e=0)
        else:
            rec.update(Scope="Unclassified", GHG_Category=f"Review — {mg}",
                       EF_val=0, kgCO2e=0, tCO2e=0)
        out.append(rec)
    return pd.DataFrame(out)


def proc_do(path):
    df = pd.read_excel(path)
    out = []
    for _, r in df.iterrows():
        qty   = sf(r.get("LFIMG",0))
        route = str(r.get("ROUTE",""))
        code  = route.replace("MY-","")
        dist  = ROUTES.get(code, 200)
        rec   = dict(
            Source="SAP DO",
            Doc_Ref=f"{r.get('VBELN','')}/{r.get('POSNR','')}",
            Date=str(r.get("LFDAT","")),
            Month="", Cost_Centre="",
            Plant=str(r.get("WERKS","")),
            Vendor=str(r.get("NAME1","")),
            Description=str(r.get("ARKTX","")),
            Material=str(r.get("MATNR","")),
            Qty=qty, Unit=str(r.get("VRKME","")),
            Spend_MYR=0, Currency="MYR",
            Route=route, Est_Dist_km=dist,
        )
        kg11=qty*EF["steel_t"]; kg9=qty*dist*EF["transport_tkm"]
        r1=dict(rec); r1.update(Scope="Scope 3 Cat 11",
            GHG_Category="Processing/Use of Sold Steel Products",
            EF_val=EF["steel_t"], kgCO2e=round(kg11,2), tCO2e=round(kg11/1000,4))
        r2=dict(rec); r2.update(Scope="Scope 3 Cat 9",
            GHG_Category="Downstream Transportation — Outbound Delivery",
            EF_val=EF["transport_tkm"], kgCO2e=round(kg9,2), tCO2e=round(kg9/1000,4))
        out.extend([r1,r2])
    return pd.DataFrame(out)


def proc_comm(path):
    df = pd.read_excel(path)
    out = []
    for _, r in df.iterrows():
        mode  = str(r.get("Transport Mode","")).strip().upper()
        dw    = sf(r.get("Days/Week",5))
        dist  = sf(r.get("One-Way Distance (km)",0))
        occ   = sf(r.get("Carpool Occupants",1)) or 1
        wfh   = sf(r.get("WFH Days",0))
        eff   = max(dw-wfh,0)
        ann   = eff*WORKING_WEEKS*dist*2
        if "MOTORCYCLE" in mode:
            ef_k=EF["motorcycle_km"]; cat="Employee Commuting — Motorcycle"
        elif "CARPOOL" in mode:
            ef_k=EF["car_km"]/occ;   cat=f"Employee Commuting — Carpool ({int(occ)} pax)"
        elif "BUS" in mode:
            ef_k=EF["bus_pax_km"];    cat="Employee Commuting — Factory Bus"
        else:
            ef_k=EF["car_km"];        cat="Employee Commuting — Private Car"
        kg=ann*ef_k
        out.append(dict(
            Source="Employee Commuting",
            Doc_Ref=str(r.get("Employee ID","")),
            Date=f"FY{REPORTING_YEAR}", Month="",
            Cost_Centre=str(r.get("Department","")), Plant="",
            Vendor=str(r.get("Employee Type","")),
            Description=f"{str(r.get('Transport Mode',''))} — {dist} km one-way",
            Spend_MYR=0, Currency="MYR",
            Scope="Scope 3 Cat 7", GHG_Category=cat,
            Qty=round(ann,2), Unit="km/year",
            EF_val=round(ef_k,4), kgCO2e=round(kg,2), tCO2e=round(kg/1000,4),
        ))
    return pd.DataFrame(out)


# ── EXCEL HELPERS ───────────────────────────────────────────────────────────────

def hf(h): return PatternFill("solid", fgColor=h)
THIN = Side(style="thin", color="D1D5DB")
MED  = Side(style="medium", color="9CA3AF")
BDR  = Border(left=THIN,right=THIN,top=THIN,bottom=THIN)
BDR2 = Border(left=MED,right=MED,top=MED,bottom=MED)

def hrow(ws, row, cols, bg, fg="FFFFFF", ht=26):
    ws.row_dimensions[row].height = ht
    for c,label in enumerate(cols,1):
        cl = ws.cell(row=row,column=c,value=label)
        cl.fill=hf(bg); cl.font=Font(bold=True,size=10,color=fg)
        cl.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        cl.border=BDR

def aw(ws,mn=8,mx=45):
    for col in ws.columns:
        w=max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width=min(max(w+2,mn),mx)

def data_sheet(wb, title, df, hbg, rbg):
    ws=wb.create_sheet(title=title)
    if df.empty: ws["A1"]=f"No records for {title}"; return
    cols=list(df.columns)
    hrow(ws,1,cols,hbg)
    rfill=hf(rbg)
    for ri,(_,row) in enumerate(df.iterrows(),2):
        ws.row_dimensions[ri].height=16
        for ci,col in enumerate(cols,1):
            v=row[col]
            cl=ws.cell(row=ri,column=ci,value=v)
            if ri%2==0: cl.fill=rfill
            cl.border=BDR
            cl.alignment=Alignment(vertical="center")
            if col in("kgCO2e","tCO2e","Spend_MYR","Qty") and isinstance(v,(int,float)):
                cl.number_format="#,##0.000"
    aw(ws); ws.freeze_panes="A2"; ws.auto_filter.ref=ws.dimensions


# ── SUPPLIER ANALYSIS SHEET ─────────────────────────────────────────────────────

def supplier_sheet(wb, po_df):
    ws=wb.create_sheet(title="🏭 Supplier Analysis")

    # Build summary
    sup=po_df.groupby("Vendor").agg(
        Spend_MYR=("Spend_MYR","sum"),
        tCO2e_S3=("tCO2e","sum"),
        PO_Count=("Doc_Ref", lambda x: len(set(str(v).split("/")[0].strip() for v in x))),
        Line_Items=("Doc_Ref","count"),
        Scope_1_tCO2e=("tCO2e", lambda x: x[po_df.loc[x.index,"Scope"].isin(["Scope 1"])].sum()),
        Scope_3_tCO2e=("tCO2e", lambda x: x[po_df.loc[x.index,"Scope"].str.startswith("Scope 3")].sum()),
        Mat_Groups=("Material_Group", lambda x: ", ".join(sorted(set(x)))),
        Top_Materials=("Description", lambda x: " / ".join(list(dict.fromkeys(x))[:3])),
    ).reset_index().sort_values("Spend_MYR",ascending=False).reset_index(drop=True)

    grand=sup["Spend_MYR"].sum()
    grand_co2=sup["tCO2e_S3"].sum()
    sup.insert(0,"Rank",range(1,len(sup)+1))
    sup["Pct_Spend"]=(sup["Spend_MYR"]/grand*100).round(2)
    sup["Cum_Pct"]=sup["Pct_Spend"].cumsum().round(2)
    sup["Pct_CO2"]=(sup["tCO2e_S3"]/grand_co2*100).round(2) if grand_co2 else 0

    def priority(cum):
        if cum<=50: return "🔴 Priority 1 — Engage Immediately"
        if cum<=80: return "🟡 Priority 2 — Engage Next"
        return "🟢 Priority 3 — Annual Monitoring"
    sup["Priority"]=sup["Cum_Pct"].apply(priority)

    # ── Title ──
    ws.merge_cells("A1:M1")
    c=ws["A1"]
    c.value="Supplier Spend Analysis — Scope 3 Cat 1 Data Collection Priority Ranking"
    c.fill=hf("0F172A"); c.font=Font(bold=True,size=16,color="F1F5F9")
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=40

    ws.merge_cells("A2:M2")
    c=ws["A2"]
    c.value=(f"Aethera Steel · FY{REPORTING_YEAR} · "
             f"Total PO Spend: MYR {grand:,.0f} "
             f"({int(sup['Pct_Spend'][:sup[sup['Cum_Pct']<=50].shape[0]].count())} suppliers = top 50% spend)  ·  "
             f"Est. Scope 3 Cat1 tCO2e from PO: {grand_co2:,.1f}  ·  FX: 1 USD = {USD_TO_MYR} MYR")
    c.fill=hf("1E293B"); c.font=Font(italic=True,size=10,color="94A3B8")
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[2].height=20

    # Legend
    ws.row_dimensions[3].height=6
    leg=[("A4:D4","🔴 Priority 1 — Top 50% of spend. Request PCF data NOW.","FEE2E2","991B1B"),
         ("E4:I4","🟡 Priority 2 — 50–80% cumulative spend. Include in supplier programme Q2.","FEF3C7","92400E"),
         ("J4:M4","🟢 Priority 3 — Bottom 20% spend. Annual monitoring sufficient.","D1FAE5","065F46")]
    for rng,txt,bg,fg in leg:
        ws.merge_cells(rng)
        s=rng.split(":")[0]
        c=ws[s]; c.value=txt; c.fill=hf(bg)
        c.font=Font(bold=True,size=9,color=fg)
        c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[4].height=22

    # Header
    hdrs=["Rank","Supplier Name","Total Spend\n(MYR)","% of\nTotal Spend",
          "Cumulative\nSpend %","Scope 1\ntCO2e (PO)","Scope 3\ntCO2e (PO)",
          "% of Total\nS3 tCO2e","POs","Line\nItems",
          "Material Groups Purchased","Top 3 Materials","Engagement Priority"]
    hrow(ws,5,hdrs,"0F172A",ht=40)

    rank_cols={1:"FEF08A",2:"E2E8F0",3:"FED7AA"}
    for ri,row in sup.iterrows():
        er=ri+6
        ws.row_dimensions[er].height=24
        cum=row["Cum_Pct"]
        bg="FFF1F2" if cum<=50 else ("FFFBEB" if cum<=80 else "F0FDF4")
        rk_bg=rank_cols.get(int(row["Rank"]),bg)

        vals=[row["Rank"],row["Vendor"],row["Spend_MYR"],row["Pct_Spend"],
              row["Cum_Pct"],row["Scope_1_tCO2e"],row["Scope_3_tCO2e"],
              row["Pct_CO2"],row["PO_Count"],row["Line_Items"],
              row["Mat_Groups"],row["Top_Materials"],row["Priority"]]

        fmts={3:"#,##0",4:"0.00",5:"0.00",6:"#,##0.0",7:"#,##0.0",8:"0.00",9:"#,##0",10:"#,##0"}

        for ci,v in enumerate(vals,1):
            cl=ws.cell(row=er,column=ci,value=v)
            cl.fill=hf(rk_bg if ci==1 else bg)
            cl.border=BDR
            cl.alignment=Alignment(horizontal="center" if ci in(1,4,5,8,9,10) else "left",
                                   vertical="center",wrap_text=(ci==13))
            if ci in fmts: cl.number_format=fmts[ci]
            if ci==2: cl.font=Font(bold=True,size=10)
            if ci==3: cl.font=Font(bold=True,size=10,color="1E3A5F")
            if ci==5:
                col_v="991B1B" if cum<=50 else ("92400E" if cum<=80 else "065F46")
                cl.font=Font(bold=True,color=col_v)

    # Totals
    tr=len(sup)+6
    ws.row_dimensions[tr].height=26
    tvs=["","TOTAL",grand,100.0,"",
         sup["Scope_1_tCO2e"].sum(),sup["Scope_3_tCO2e"].sum(),100.0,
         sup["PO_Count"].sum(),sup["Line_Items"].sum(),
         f"{len(sup)} suppliers","",""]
    for ci,v in enumerate(tvs,1):
        cl=ws.cell(row=tr,column=ci,value=v)
        cl.fill=hf("FEF9C3"); cl.font=Font(bold=True,size=11)
        cl.alignment=Alignment(horizontal="center",vertical="center")
        cl.border=BDR2
        fmts2={3:"#,##0",6:"#,##0.0",7:"#,##0.0",9:"#,##0",10:"#,##0"}
        if ci in fmts2: cl.number_format=fmts2[ci]

    # Bar chart
    chart=BarChart()
    chart.type="col"; chart.title="Supplier Spend Ranking (MYR)"
    chart.y_axis.title="Total Spend (MYR)"
    chart.style=10; chart.width=28; chart.height=16
    dr=Reference(ws,min_col=3,min_row=5,max_row=5+len(sup))
    cr=Reference(ws,min_col=2,min_row=6,max_row=5+len(sup))
    chart.add_data(dr,titles_from_data=True)
    chart.set_categories(cr)
    for i,(_,row) in enumerate(sup.iterrows()):
        c_hex="C0392B" if row["Cum_Pct"]<=50 else ("D97706" if row["Cum_Pct"]<=80 else "16A34A")
        pt=DataPoint(idx=i); pt.graphicalProperties.solidFill=c_hex
        chart.series[0].dPt.append(pt)
    ws.add_chart(chart,f"B{tr+3}")

    # Material group breakdown
    sub_r=tr+22
    ws.merge_cells(f"A{sub_r}:M{sub_r}")
    c=ws.cell(row=sub_r,column=1,
        value="Material Group × Supplier Spend Breakdown (MYR) — Prioritise which material categories to collect PCF data for")
    c.fill=hf("0F172A"); c.font=Font(bold=True,size=12,color="FFFFFF")
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[sub_r].height=30

    mg_piv=po_df.pivot_table(index="Vendor",columns="Material_Group",
                              values="Spend_MYR",aggfunc="sum",fill_value=0)
    mg_piv=mg_piv.reindex(sup["Vendor"].values)
    mg_piv["TOTAL"]=mg_piv.sum(axis=1)
    mg_piv=mg_piv.sort_values("TOTAL",ascending=False)

    sh2=sub_r+1
    mg_hdrs=["Supplier"]+list(mg_piv.columns)
    hrow(ws,sh2,mg_hdrs,"1E293B",ht=28)

    # Colour code matgroup columns
    scope_col={"ENERGY":"1A4731","FUEL-RED":"1A4731","RAW-IRON":"4C1D95",
               "SCRAP":"5B21B6","FLUX":"6D28D9","ALLOY":"7C3AED",
               "IND-GAS":"8B5CF6","PACK":"374151","TOTAL":"78350F"}
    for ci,col in enumerate(mg_piv.columns,2):
        bg2=scope_col.get(col,"374151")
        ws.cell(row=sh2,column=ci).fill=hf(bg2)

    for ri2,(vendor,row) in enumerate(mg_piv.iterrows()):
        er2=sh2+1+ri2
        ws.row_dimensions[er2].height=20
        bg3="F8FAFC" if ri2%2==0 else "EFF6FF"
        cl=ws.cell(row=er2,column=1,value=vendor)
        cl.font=Font(bold=True,size=10); cl.border=BDR; cl.fill=hf(bg3)
        for ci,col in enumerate(mg_piv.columns,2):
            v=row[col]
            cl=ws.cell(row=er2,column=ci,value=round(v,0) if v else None)
            cl.border=BDR; cl.alignment=Alignment(horizontal="right",vertical="center")
            cl.number_format="#,##0"
            if col=="TOTAL":
                cl.font=Font(bold=True,color="1E3A5F"); cl.fill=hf("DBEAFE")
            elif v>0:
                cl.fill=hf(bg3)
            else:
                cl.fill=hf("F9FAFB"); cl.value="—"

    # Widths
    ws.column_dimensions["A"].width=36
    for ci in range(2,len(hdrs)+1):
        ws.column_dimensions[get_column_letter(ci)].width=16
    ws.freeze_panes="B6"


# ── SUMMARY SHEET ───────────────────────────────────────────────────────────────

def summary_sheet(wb, all_data):
    ws=wb.create_sheet(title="📊 GHG Summary",index=0)

    ws.merge_cells("A1:J1")
    c=ws["A1"]; c.value="GHG Emission Inventory — Aethera Steel (AETH) FY2025"
    c.fill=hf("0F172A"); c.font=Font(bold=True,size=18,color="F8FAFC")
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[1].height=44

    ws.merge_cells("A2:J2")
    c=ws["A2"]
    c.value="GHG Protocol (ISO 14064-1)  ·  DEFRA 2024 / IEA MY Grid / IPCC AR6 / World Steel 2023  ·  Sources: SAP PO · SAP DO · SAP FI · Employee Survey"
    c.fill=hf("1E293B"); c.font=Font(italic=True,size=10,color="94A3B8")
    c.alignment=Alignment(horizontal="center",vertical="center")
    ws.row_dimensions[2].height=20
    ws.row_dimensions[3].height=10

    scopes=[
        ("Scope 1","Direct Emissions (stationary + mobile combustion)","1A4731","D1FAE5"),
        ("Scope 2","Purchased Electricity (TNB — market-based)","1E3A5F","DBEAFE"),
        ("Scope 3 Cat 1","Purchased Goods & Services (iron ore, scrap, flux, alloys, gases)","3B0764","EDE9FE"),
        ("Scope 3 Cat 5","Waste Generated in Operations (water / wastewater — Air Selangor)","4C1D95","F5F3FF"),
        ("Scope 3 Cat 7","Employee Commuting (car, motorcycle, carpool, factory bus)","5B21B6","F5F3FF"),
        ("Scope 3 Cat 9","Downstream Transportation (outbound steel deliveries)","6D28D9","EDE9FE"),
        ("Scope 3 Cat 11","Processing & Use of Sold Steel Products","7C3AED","EDE9FE"),
        ("Unclassified","Items requiring review","374151","F3F4F6"),
    ]

    hrow(ws,4,["Scope","GHG Protocol Category","tCO2e","% of Total",
               "Cumulative %","Record Count","Data Quality"],"0F172A",ht=28)

    grand=all_data["tCO2e"].sum(); cum=0; dr=5
    for scope,cat,hc,rc in scopes:
        sub=all_data[all_data["Scope"]==scope]
        tot=sub["tCO2e"].sum(); pct=tot/grand*100 if grand else 0; cum+=pct
        dq="Metered/Measured" if scope in("Scope 1","Scope 2") else \
           "Activity-based" if scope.startswith("Scope 3") else "—"
        ws.row_dimensions[dr].height=22
        for ci,v in enumerate([scope,cat,round(tot,3),round(pct,2),round(cum,2),len(sub),dq],1):
            cl=ws.cell(row=dr,column=ci,value=v)
            cl.fill=hf(rc); cl.border=BDR
            cl.alignment=Alignment(horizontal="center" if ci in(3,4,5,6) else "left",vertical="center")
            if ci==3: cl.number_format="#,##0.000"
            if ci in(4,5): cl.number_format="0.00"
            if ci==1: cl.font=Font(bold=True,size=10,color=hc)
        dr+=1

    # Grand total
    ws.row_dimensions[dr].height=26
    for ci,v in enumerate(["GRAND TOTAL","All Scopes",round(grand,3),100.0,"",len(all_data),""],1):
        cl=ws.cell(row=dr,column=ci,value=v)
        cl.fill=hf("FEF9C3"); cl.font=Font(bold=True,size=12); cl.border=BDR2
        cl.alignment=Alignment(horizontal="center",vertical="center")
        if ci==3: cl.number_format="#,##0.000"

    # Source breakdown
    dr+=2
    ws.cell(row=dr,column=1,value="Breakdown by Source File").font=Font(bold=True,size=13)
    dr+=1
    hrow(ws,dr,["Source File","Records","tCO2e","% of Total","Primary Scope"],"1E293B",ht=22)
    dr+=1
    for src,grp in all_data.groupby("Source"):
        t=grp["tCO2e"].sum(); p=t/grand*100 if grand else 0
        sc=grp["Scope"].value_counts().index[0] if len(grp) else "—"
        for ci,v in enumerate([src,len(grp),round(t,3),round(p,2),sc],1):
            cl=ws.cell(row=dr,column=ci,value=v); cl.border=BDR
            if ci==3: cl.number_format="#,##0.000"
            if ci==4: cl.number_format="0.00"
        dr+=1

    # EF table
    dr+=2
    ws.cell(row=dr,column=1,value="Emission Factors Used").font=Font(bold=True,size=13)
    dr+=1
    hrow(ws,dr,["Emission Source","EF Value","Unit","Reference Standard"],"1E293B",ht=22)
    dr+=1
    eft=[("Electricity (MY Grid)",0.585,"kgCO2e/kWh","IEA 2024 Malaysia"),
         ("Natural Gas",56.1,"kgCO2e/MMBTU","IPCC AR6"),
         ("Diesel",2.686,"kgCO2e/litre","DEFRA 2024"),
         ("Fuel Oil",3.179,"kgCO2e/litre","DEFRA 2024"),
         ("Metallurgical Coke",3.96,"tCO2e/tonne","IPCC AR6 / GHG Protocol"),
         ("Iron Ore",0.025,"tCO2e/tonne","World Steel 2023"),
         ("Quicklime",0.785,"tCO2e/tonne","IPCC AR6"),
         ("Dolomite",0.477,"tCO2e/tonne","IPCC AR6"),
         ("Ferro Silicon 75%",4.0,"tCO2e/tonne","World Steel 2023"),
         ("HCFeMn",1.8,"tCO2e/tonne","World Steel 2023"),
         ("Steel Products (sold)",1.83,"tCO2e/tonne","World Steel avg 2023"),
         ("Private Car",0.170,"kgCO2e/km","DEFRA 2024"),
         ("Motorcycle",0.103,"kgCO2e/km","DEFRA 2024"),
         ("Factory Bus",0.089,"kgCO2e/pax-km","DEFRA 2024"),
         ("Road Freight",0.096,"kgCO2e/tonne-km","DEFRA 2024"),
         ("Water Treatment",0.000344,"kgCO2e/litre","IPCC AR6")]
    for i,(s,v,u,ref) in enumerate(eft):
        bg="F8FAFC" if i%2==0 else "EFF6FF"
        for ci,val in enumerate([s,v,u,ref],1):
            cl=ws.cell(row=dr,column=ci,value=val)
            cl.fill=hf(bg); cl.border=BDR
        dr+=1

    for col,w in zip("ABCDEFGHIJ",[28,42,14,12,14,12,26,8,8,8]):
        ws.column_dimensions[col].width=w
    ws.freeze_panes="A5"


# ── MAIN ────────────────────────────────────────────────────────────────────────

def main():
    print("\n🌿 Aethera GHG Classifier v2\n")
    fi=proc_fi(find("SAP_FI_Utility_Export_Realistic.xlsx"))
    po=proc_po(find("SAP_Steel_Procurement_Actual_Style_Dummy.xlsx"))
    do=proc_do(find("SAP_Steel_DO_Dummy_100DO.xlsx"))
    cm=proc_comm(find("Aethera_Scope3_Category7_Dummy_Data.xlsx"))

    all_d=pd.concat([fi,po,do,cm],ignore_index=True,sort=False)
    for col in ["tCO2e","kgCO2e","Spend_MYR"]:
        if col in all_d.columns:
            all_d[col]=pd.to_numeric(all_d[col],errors="coerce").fillna(0)

    print("═"*65)
    print("  GHG SUMMARY (tCO2e)"); print("─"*65)
    tots=all_d.groupby("Scope")["tCO2e"].sum().sort_values(ascending=False)
    g=tots.sum()
    for s,v in tots.items(): print(f"  {s:<26} {v:>12,.3f}  ({v/g*100:.1f}%)")
    print("─"*65); print(f"  {'TOTAL':<26} {g:>12,.3f}")
    print("═"*65)

    print("\n  SUPPLIER SPEND RANKING"); print("─"*65)
    sup=po.groupby("Vendor")["Spend_MYR"].sum().sort_values(ascending=False)
    grand=sup.sum(); cum=0
    for i,(v,s) in enumerate(sup.items(),1):
        pct=s/grand*100; cum+=pct
        fl="🔴" if cum<=50 else ("🟡" if cum<=80 else "🟢")
        print(f"  {i:>2}. {fl} {v:<36} MYR {s:>11,.0f}  ({pct:.1f}%  cum {cum:.1f}%)")
    print("═"*65)

    wb=Workbook(); wb.remove(wb.active)
    summary_sheet(wb, all_d)
    supplier_sheet(wb, po)

    sheets=[
        ("Scope 1 — Direct",         "Scope 1",       "1A4731","D1FAE5"),
        ("Scope 2 — Electricity",    "Scope 2",       "1E3A5F","DBEAFE"),
        ("Scope 3 Cat1 — Purchased", "Scope 3 Cat 1", "3B0764","EDE9FE"),
        ("Scope 3 Cat5 — Waste",     "Scope 3 Cat 5", "4C1D95","F5F3FF"),
        ("Scope 3 Cat7 — Commuting", "Scope 3 Cat 7", "5B21B6","F5F3FF"),
        ("Scope 3 Cat9 — Transport", "Scope 3 Cat 9", "6D28D9","EDE9FE"),
        ("Scope 3 Cat11 — Products", "Scope 3 Cat 11","7C3AED","EDE9FE"),
    ]
    for title,scope,hbg,rbg in sheets:
        data_sheet(wb,title,all_d[all_d["Scope"]==scope],hbg,rbg)
    unc=all_d[all_d["Scope"]=="Unclassified"]
    if not unc.empty: data_sheet(wb,"⚠ Unclassified",unc,"374151","F3F4F6")

    os.makedirs("/mnt/user-data/outputs",exist_ok=True)
    wb.save(OUTPUT)
    print(f"\n✅ Saved: {OUTPUT}")

if __name__=="__main__":
    main()
