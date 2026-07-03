import os
import io
import pickle
import uuid
import json
import tempfile
import pandas as pd
from flask import (
    Flask, request, redirect, url_for, render_template,
    session, jsonify, send_file
)
from werkzeug.utils import secure_filename
from dotenv import load_dotenv
import psycopg2
from psycopg2.extras import execute_values

load_dotenv()

app = Flask(__name__)
app.secret_key = os.urandom(24)

DATABASE_URL = os.environ.get("DATABASE_URL")


def _get_db():
    return psycopg2.connect(DATABASE_URL)


def _init_db():
    with _get_db() as conn:
        with conn.cursor() as cur:
            cur.execute("""
                CREATE TABLE IF NOT EXISTS ghg_emissions (
                    id          SERIAL PRIMARY KEY,
                    session_id  TEXT,
                    uploaded_at TIMESTAMPTZ DEFAULT NOW(),
                    source      TEXT,
                    doc_ref     TEXT,
                    date        TEXT,
                    vendor      TEXT,
                    description TEXT,
                    scope       TEXT,
                    ghg_category TEXT,
                    qty         NUMERIC,
                    unit        TEXT,
                    ef_val      NUMERIC,
                    kg_co2e     NUMERIC,
                    t_co2e      NUMERIC,
                    spend_myr   NUMERIC,
                    plant       TEXT
                )
            """)
        conn.commit()


def _save_to_db(sid, df):
    rows = [
        (
            sid,
            r.get("Source"), r.get("Doc_Ref"), str(r.get("Date") or ""),
            r.get("Vendor"), r.get("Description"), r.get("Scope"),
            r.get("GHG_Category"),
            float(r["Qty"]) if r.get("Qty") is not None else None,
            r.get("Unit"),
            float(r["EF_val"]) if r.get("EF_val") is not None else None,
            float(r["kgCO2e"]) if r.get("kgCO2e") is not None else None,
            float(r["tCO2e"]) if r.get("tCO2e") is not None else None,
            float(r["Spend_MYR"]) if r.get("Spend_MYR") is not None else None,
            r.get("Plant"),
        )
        for _, r in df.iterrows()
    ]
    with _get_db() as conn:
        with conn.cursor() as cur:
            execute_values(cur, """
                INSERT INTO ghg_emissions
                    (session_id, source, doc_ref, date, vendor, description,
                     scope, ghg_category, qty, unit, ef_val, kg_co2e, t_co2e,
                     spend_myr, plant)
                VALUES %s
            """, rows)
        conn.commit()


_init_db()

UPLOAD_FOLDER = os.path.join(os.path.dirname(__file__), "uploads")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# ── CONFIG ────────────────────────────────────────────────────────────────────
REPORTING_YEAR = 2025
WORKING_WEEKS  = 48
USD_TO_MYR     = 4.47
MYR_PER_KWH    = 0.509
MYR_PER_DIESEL = 2.30
MYR_PER_FOIL   = 2.80
MYR_PER_MMBTU  = 21.0

EF = {
    "diesel_litre":  2.686,
    "fuel_oil_litre": 3.179,
    "nat_gas_mmbtu": 56.1,
    "met_coke_t":    3960.0,
    "electricity_kwh": 0.585,
    "iron_ore_t":    25.0,
    "scrap_t":       1.0,
    "quicklime_t":   785.0,
    "dolomite_t":    477.0,
    "ferro_si_t":    4000.0,
    "ferro_mn_t":    1800.0,
    "liq_oxy_m3":    0.26,
    "nitrogen_m3":   0.10,
    "transport_tkm": 0.096,
    "car_km":        0.170,
    "motorcycle_km": 0.103,
    "bus_pax_km":    0.089,
    "steel_t":       1830.0,
    "water_litre":   0.000344,
}

ROUTES = {
    "KUL": 50, "PNG": 320, "JHR": 360, "PRK": 200,
    "SGR": 40, "SBH": 1500, "SWK": 800, "SLG": 90,
}

OUTPUT_COLS = [
    "Source", "Doc_Ref", "Date", "Vendor", "Description",
    "Scope", "GHG_Category", "Qty", "Unit", "EF_val",
    "kgCO2e", "tCO2e", "Spend_MYR", "Plant",
]

# ── HELPERS ───────────────────────────────────────────────────────────────────
def _empty_row():
    return {c: None for c in OUTPUT_COLS}


def _to_float(val, default=0.0):
    try:
        return float(val)
    except (TypeError, ValueError):
        return default


def _col(df, *candidates):
    """Return the first matching column name (case-insensitive) or None."""
    lower = {c.lower(): c for c in df.columns}
    for cand in candidates:
        if cand.lower() in lower:
            return lower[cand.lower()]
    return None


def _get(row, *candidates):
    for cand in candidates:
        for col in row.index:
            if col.lower() == cand.lower():
                return row[col]
    return None


# ── SAP FI PROCESSOR ──────────────────────────────────────────────────────────
def process_fi(path):
    try:
        df = pd.read_excel(path, dtype=str)
    except Exception as e:
        return [], f"FI read error: {e}"

    df.columns = df.columns.str.strip()
    records = []

    gl_col   = _col(df, "GL_ACCOUNT_TEXT", "GL Account Text", "GL_ACCOUNT")
    amt_col  = _col(df, "WRBTR", "Amount", "Net Amount")
    ref_col  = _col(df, "BELNR", "Doc No", "Document Number")
    date_col = _col(df, "BUDAT", "Posting Date", "Date")
    name_col = _col(df, "NAME1", "Vendor Name", "Vendor")
    plant_col = _col(df, "WERKS", "Plant")
    cost_col  = _col(df, "KOSTL", "Cost Centre")

    if gl_col is None or amt_col is None:
        return [], "FI: missing GL_ACCOUNT_TEXT or WRBTR columns"

    for _, row in df.iterrows():
        gl   = str(_get(row, gl_col) or "").upper()
        amt  = _to_float(_get(row, amt_col))
        if amt == 0:
            continue

        r = _empty_row()
        r["Source"]    = "SAP FI"
        r["Doc_Ref"]   = _get(row, ref_col) if ref_col else None
        r["Date"]      = _get(row, date_col) if date_col else None
        r["Vendor"]    = _get(row, name_col) if name_col else None
        r["Plant"]     = _get(row, plant_col) if plant_col else None
        r["Spend_MYR"] = amt

        if "ELECTRICITY" in gl:
            kwh = amt / MYR_PER_KWH
            r.update(Description="Electricity", Scope="Scope 2",
                     GHG_Category="Scope 2 — Electricity",
                     Qty=round(kwh, 2), Unit="kWh",
                     EF_val=EF["electricity_kwh"],
                     kgCO2e=round(kwh * EF["electricity_kwh"], 2))
        elif "NATURAL GAS" in gl or "NGAS" in gl:
            mmbtu = amt / MYR_PER_MMBTU
            r.update(Description="Natural Gas", Scope="Scope 1",
                     GHG_Category="Scope 1 — Natural Gas",
                     Qty=round(mmbtu, 2), Unit="MMBTU",
                     EF_val=EF["nat_gas_mmbtu"],
                     kgCO2e=round(mmbtu * EF["nat_gas_mmbtu"], 2))
        elif "DIESEL" in gl:
            litres = amt / MYR_PER_DIESEL
            r.update(Description="Diesel", Scope="Scope 1",
                     GHG_Category="Scope 1 — Diesel",
                     Qty=round(litres, 2), Unit="Litres",
                     EF_val=EF["diesel_litre"],
                     kgCO2e=round(litres * EF["diesel_litre"], 2))
        elif "FUEL OIL" in gl:
            litres = amt / MYR_PER_FOIL
            r.update(Description="Fuel Oil", Scope="Scope 1",
                     GHG_Category="Scope 1 — Fuel Oil",
                     Qty=round(litres, 2), Unit="Litres",
                     EF_val=EF["fuel_oil_litre"],
                     kgCO2e=round(litres * EF["fuel_oil_litre"], 2))
        elif "WATER" in gl:
            litres = amt / 0.003
            r.update(Description="Water", Scope="Scope 3",
                     GHG_Category="Scope 3 Cat 5 — Waste / Water",
                     Qty=round(litres, 2), Unit="Litres",
                     EF_val=EF["water_litre"],
                     kgCO2e=round(litres * EF["water_litre"], 2))
        else:
            continue

        r["tCO2e"] = round(r["kgCO2e"] / 1000, 4)
        records.append(r)

    return records, None


# ── SAP PO PROCESSOR ──────────────────────────────────────────────────────────
_PO_RULES = [
    # (mat_group_kw, short_text_kw, scope, ghg_cat, ef_key, unit)
    ("ENERGY",   "DIESEL",       "Scope 1", "Scope 1 — Diesel",        "diesel_litre",  "Litres"),
    ("ENERGY",   "NGAS",         "Scope 1", "Scope 1 — Natural Gas",   "nat_gas_mmbtu", "MMBTU"),
    ("ENERGY",   "NATURAL GAS",  "Scope 1", "Scope 1 — Natural Gas",   "nat_gas_mmbtu", "MMBTU"),
    ("FUEL-RED", None,           "Scope 1", "Scope 1 — Met Coke",      "met_coke_t",    "Tonnes"),
    ("RAW-IRON", None,           "Scope 3", "Scope 3 Cat 1 — Purchased Goods", "iron_ore_t", "Tonnes"),
    ("SCRAP",    None,           "Scope 3", "Scope 3 Cat 1 — Purchased Goods", "scrap_t",    "Tonnes"),
    ("FLUX",     "LIME",         "Scope 3", "Scope 3 Cat 1 — Purchased Goods", "quicklime_t","Tonnes"),
    ("FLUX",     "QUICKLIME",    "Scope 3", "Scope 3 Cat 1 — Purchased Goods", "quicklime_t","Tonnes"),
    ("FLUX",     "DOLO",         "Scope 3", "Scope 3 Cat 1 — Purchased Goods", "dolomite_t", "Tonnes"),
    ("FLUX",     "DOLOMITE",     "Scope 3", "Scope 3 Cat 1 — Purchased Goods", "dolomite_t", "Tonnes"),
    ("ALLOY",    "FESI",         "Scope 3", "Scope 3 Cat 1 — Purchased Goods", "ferro_si_t", "Tonnes"),
    ("ALLOY",    "SILICON",      "Scope 3", "Scope 3 Cat 1 — Purchased Goods", "ferro_si_t", "Tonnes"),
    ("ALLOY",    "FEMN",         "Scope 3", "Scope 3 Cat 1 — Purchased Goods", "ferro_mn_t", "Tonnes"),
    ("ALLOY",    "MANGANESE",    "Scope 3", "Scope 3 Cat 1 — Purchased Goods", "ferro_mn_t", "Tonnes"),
    ("IND-GAS",  "OXY",         "Scope 3", "Scope 3 Cat 1 — Purchased Goods", "liq_oxy_m3", "m3"),
    ("IND-GAS",  "OXYGEN",      "Scope 3", "Scope 3 Cat 1 — Purchased Goods", "liq_oxy_m3", "m3"),
    ("IND-GAS",  "NIT",         "Scope 3", "Scope 3 Cat 1 — Purchased Goods", "nitrogen_m3","m3"),
    ("IND-GAS",  "NITROGEN",    "Scope 3", "Scope 3 Cat 1 — Purchased Goods", "nitrogen_m3","m3"),
    ("PACK",     None,          "Scope 3", "Scope 3 Cat 1 — Purchased Goods", None,         "Units"),
]


def process_po(path):
    try:
        df = pd.read_excel(path, dtype=str)
    except Exception as e:
        return [], f"PO read error: {e}"

    df.columns = df.columns.str.strip()
    records = []

    mat_col  = _col(df, "Material Group", "MaterialGroup", "Mat. Group")
    txt_col  = _col(df, "Short Text", "ShortText", "Material Description")
    qty_col  = _col(df, "PO Quantity", "Quantity", "Order Quantity")
    val_col  = _col(df, "Net Order Value", "Net Value", "Order Value")
    cur_col  = _col(df, "Currency", "Curr.", "Crcy")
    vnd_col  = _col(df, "Vendor Name", "Vendor", "Supplier")
    ref_col  = _col(df, "Purch.Doc.", "PO Number", "Doc. No")
    date_col = _col(df, "Doc. Date", "Document Date", "PO Date")
    plt_col  = _col(df, "Plant")

    if mat_col is None:
        return [], "PO: missing Material Group column"

    for _, row in df.iterrows():
        mat_grp = str(_get(row, mat_col) or "").upper().strip()
        short   = str(_get(row, txt_col) or "").upper().strip() if txt_col else ""
        qty     = _to_float(_get(row, qty_col) if qty_col else 0)
        val     = _to_float(_get(row, val_col) if val_col else 0)
        cur     = str(_get(row, cur_col) or "MYR").upper().strip() if cur_col else "MYR"
        spend   = val * USD_TO_MYR if cur == "USD" else val

        matched = False
        for (mg_kw, st_kw, scope, ghg_cat, ef_key, unit) in _PO_RULES:
            if mg_kw not in mat_grp:
                continue
            if st_kw and st_kw not in short:
                continue
            matched = True
            ef_val  = EF.get(ef_key, 0) if ef_key else 0
            kgco2e  = round(qty * ef_val, 2)

            r = _empty_row()
            r.update(
                Source="SAP PO",
                Doc_Ref=_get(row, ref_col) if ref_col else None,
                Date=_get(row, date_col) if date_col else None,
                Vendor=_get(row, vnd_col) if vnd_col else None,
                Description=short or mat_grp,
                Scope=scope,
                GHG_Category=ghg_cat,
                Qty=qty,
                Unit=unit,
                EF_val=ef_val,
                kgCO2e=kgco2e,
                tCO2e=round(kgco2e / 1000, 4),
                Spend_MYR=round(spend, 2),
                Plant=_get(row, plt_col) if plt_col else None,
            )
            records.append(r)
            break

        if not matched and mat_grp:
            r = _empty_row()
            r.update(
                Source="SAP PO",
                Doc_Ref=_get(row, ref_col) if ref_col else None,
                Date=_get(row, date_col) if date_col else None,
                Vendor=_get(row, vnd_col) if vnd_col else None,
                Description=short or mat_grp,
                Scope="Unclassified",
                GHG_Category="Unclassified",
                Qty=qty,
                Unit="—",
                EF_val=0,
                kgCO2e=0,
                tCO2e=0,
                Spend_MYR=round(spend, 2),
                Plant=_get(row, plt_col) if plt_col else None,
            )
            records.append(r)

    return records, None


# ── SAP DO PROCESSOR ──────────────────────────────────────────────────────────
def process_do(path):
    try:
        df = pd.read_excel(path, dtype=str)
    except Exception as e:
        return [], f"DO read error: {e}"

    df.columns = df.columns.str.strip()
    records = []

    qty_col   = _col(df, "LFIMG", "Delivery Qty", "Quantity")
    rte_col   = _col(df, "ROUTE", "Route", "Shipping Route")
    ref_col   = _col(df, "VBELN", "Delivery", "Delivery No")
    itm_col   = _col(df, "POSNR", "Item")
    mat_col   = _col(df, "ARKTX", "Material Description", "Description")
    cust_col  = _col(df, "NAME1", "Customer Name", "Customer")
    date_col  = _col(df, "LFDAT", "Delivery Date", "Date")
    plt_col   = _col(df, "WERKS", "Plant")

    if qty_col is None:
        return [], "DO: missing LFIMG / quantity column"

    for _, row in df.iterrows():
        qty   = _to_float(_get(row, qty_col) if qty_col else 0)
        route = str(_get(row, rte_col) or "").upper().strip() if rte_col else ""
        suffix = route[-3:] if len(route) >= 3 else route
        dist  = ROUTES.get(suffix, 200)

        base = dict(
            Source="SAP DO",
            Doc_Ref=_get(row, ref_col) if ref_col else None,
            Date=_get(row, date_col) if date_col else None,
            Vendor=_get(row, cust_col) if cust_col else None,
            Description=_get(row, mat_col) if mat_col else None,
            Plant=_get(row, plt_col) if plt_col else None,
            Spend_MYR=None,
        )

        # Cat 11 — Use of sold products (steel)
        kg11 = round(qty * EF["steel_t"], 2)
        r11  = _empty_row()
        r11.update(**base)
        r11.update(
            Scope="Scope 3",
            GHG_Category="Scope 3 Cat 11 — Use of Sold Products",
            Qty=qty, Unit="Tonnes",
            EF_val=EF["steel_t"],
            kgCO2e=kg11, tCO2e=round(kg11 / 1000, 4),
        )
        records.append(r11)

        # Cat 9 — Downstream transport
        tkm   = qty * dist
        kg9   = round(tkm * EF["transport_tkm"], 2)
        r9    = _empty_row()
        r9.update(**base)
        r9.update(
            Scope="Scope 3",
            GHG_Category="Scope 3 Cat 9 — Downstream Transport",
            Qty=round(tkm, 2), Unit="tonne-km",
            EF_val=EF["transport_tkm"],
            kgCO2e=kg9, tCO2e=round(kg9 / 1000, 4),
        )
        records.append(r9)

    return records, None


# ── COMMUTING PROCESSOR ───────────────────────────────────────────────────────
def process_commuting(path):
    try:
        df = pd.read_excel(path, dtype=str)
    except Exception as e:
        return [], f"Commuting read error: {e}"

    df.columns = df.columns.str.strip()
    records = []

    id_col   = _col(df, "Employee ID", "EmpID", "Staff ID")
    dept_col = _col(df, "Department", "Dept")
    mode_col = _col(df, "Transport Mode", "Mode", "Transport")
    days_col = _col(df, "Days/Week", "Working Days/Week", "Days Per Week")
    dist_col = _col(df, "One-Way Distance (km)", "Distance (km)", "One Way Distance")
    pool_col = _col(df, "Carpool Occupants", "Occupants", "Carpool")
    wfh_col  = _col(df, "WFH Days", "WFH Days/Week", "Work From Home Days")
    type_col = _col(df, "Employee Type", "Emp Type")

    if mode_col is None or dist_col is None:
        return [], "Commuting: missing Transport Mode or Distance columns"

    for _, row in df.iterrows():
        mode    = str(_get(row, mode_col) or "").upper().strip()
        dist    = _to_float(_get(row, dist_col) if dist_col else 0)
        days    = _to_float(_get(row, days_col) if days_col else 5)
        wfh     = _to_float(_get(row, wfh_col) if wfh_col else 0)
        pool    = _to_float(_get(row, pool_col) if pool_col else 1) or 1

        annual_km = (days - wfh) * WORKING_WEEKS * dist * 2

        if "MOTORCYCLE" in mode:
            ef_val = EF["motorcycle_km"]
            unit   = "km"
        elif "CARPOOL" in mode:
            ef_val = EF["car_km"] / pool
            unit   = "km"
        elif "BUS" in mode or "FACTORY BUS" in mode:
            ef_val = EF["bus_pax_km"]
            unit   = "km"
        else:
            ef_val = EF["car_km"]
            unit   = "km"

        kgco2e = round(annual_km * ef_val, 2)

        r = _empty_row()
        r.update(
            Source="Commuting Survey",
            Doc_Ref=_get(row, id_col) if id_col else None,
            Date=str(REPORTING_YEAR),
            Vendor=_get(row, dept_col) if dept_col else None,
            Description=f"{mode} commute",
            Scope="Scope 3",
            GHG_Category="Scope 3 Cat 7 — Employee Commuting",
            Qty=round(annual_km, 2),
            Unit=unit,
            EF_val=round(ef_val, 4),
            kgCO2e=kgco2e,
            tCO2e=round(kgco2e / 1000, 4),
            Spend_MYR=None,
            Plant=None,
        )
        records.append(r)

    return records, None


# ── SESSION STORAGE ───────────────────────────────────────────────────────────
def _session_path(sid):
    return os.path.join(tempfile.gettempdir(), f"ghg_{sid}.pkl")


def save_session(sid, df):
    with open(_session_path(sid), "wb") as f:
        pickle.dump(df, f)


def load_session(sid):
    path = _session_path(sid)
    if not os.path.exists(path):
        return None
    with open(path, "rb") as f:
        return pickle.load(f)


# ── ROUTES ────────────────────────────────────────────────────────────────────
@app.route("/")
def index():
    return render_template("index.html")


@app.route("/process", methods=["POST"])
def process():
    file_keys = {
        "sap_po":        process_po,
        "sap_fi":        process_fi,
        "sap_do":        process_do,
        "commuting":     process_commuting,
    }

    all_records = []
    errors = []

    for key, processor in file_keys.items():
        f = request.files.get(key)
        if not f or f.filename == "":
            continue
        fname = secure_filename(f.filename)
        save_path = os.path.join(UPLOAD_FOLDER, fname)
        f.save(save_path)
        recs, err = processor(save_path)
        if err:
            errors.append(err)
        else:
            all_records.extend(recs)

    if not all_records:
        msg = "No records processed."
        if errors:
            msg += " Errors: " + "; ".join(errors)
        return jsonify({"error": msg}), 400

    df = pd.DataFrame(all_records, columns=OUTPUT_COLS)
    df["kgCO2e"]   = pd.to_numeric(df["kgCO2e"],   errors="coerce").fillna(0)
    df["tCO2e"]    = pd.to_numeric(df["tCO2e"],     errors="coerce").fillna(0)
    df["Spend_MYR"] = pd.to_numeric(df["Spend_MYR"], errors="coerce").fillna(0)

    sid = str(uuid.uuid4())
    save_session(sid, df)
    try:
        _save_to_db(sid, df)
    except Exception as db_err:
        errors.append(f"DB save warning: {db_err}")
    session["ghg_sid"] = sid

    if errors:
        session["ghg_warnings"] = errors

    return redirect(url_for("dashboard"))


@app.route("/dashboard")
def dashboard():
    sid = session.get("ghg_sid")
    if not sid:
        return redirect(url_for("index"))

    df = load_session(sid)
    if df is None:
        return redirect(url_for("index"))

    warnings = session.pop("ghg_warnings", [])

    # ── Summary totals ────────────────────────────────────────────────────────
    total   = round(df["tCO2e"].sum(), 2)
    scope1  = round(df.loc[df["Scope"] == "Scope 1", "tCO2e"].sum(), 2)
    scope2  = round(df.loc[df["Scope"] == "Scope 2", "tCO2e"].sum(), 2)
    scope3  = round(df.loc[df["Scope"] == "Scope 3", "tCO2e"].sum(), 2)

    # ── Donut chart data ──────────────────────────────────────────────────────
    cat_totals = (
        df.groupby("GHG_Category")["tCO2e"]
          .sum()
          .round(2)
          .reset_index()
    )
    cat_totals = cat_totals[cat_totals["tCO2e"] > 0]

    donut_labels = cat_totals["GHG_Category"].tolist()
    donut_values = cat_totals["tCO2e"].tolist()

    colour_map = {
        "Scope 1": "#22c55e",
        "Scope 2": "#3b82f6",
        "Scope 3 Cat 1": "#a855f7",
        "Scope 3 Cat 5": "#9333ea",
        "Scope 3 Cat 7": "#7c3aed",
        "Scope 3 Cat 9": "#6d28d9",
        "Scope 3 Cat 11": "#4c1d95",
    }

    def _colour(label):
        label_u = label.upper()
        if "SCOPE 1" in label_u:
            return colour_map["Scope 1"]
        if "SCOPE 2" in label_u:
            return colour_map["Scope 2"]
        if "CAT 1" in label_u:
            return colour_map["Scope 3 Cat 1"]
        if "CAT 5" in label_u:
            return colour_map["Scope 3 Cat 5"]
        if "CAT 7" in label_u:
            return colour_map["Scope 3 Cat 7"]
        if "CAT 9" in label_u:
            return colour_map["Scope 3 Cat 9"]
        if "CAT 11" in label_u:
            return colour_map["Scope 3 Cat 11"]
        return "#94a3b8"

    donut_colours = [_colour(l) for l in donut_labels]

    # ── Top 10 suppliers ──────────────────────────────────────────────────────
    po_df = df[df["Source"] == "SAP PO"].copy()
    if not po_df.empty:
        sup = (
            po_df.groupby("Vendor")["Spend_MYR"]
                 .sum()
                 .sort_values(ascending=False)
                 .head(10)
                 .reset_index()
        )
        sup["Spend_MYR"] = sup["Spend_MYR"].round(2)
        total_spend = sup["Spend_MYR"].sum()
        sup["cum_pct"] = (sup["Spend_MYR"].cumsum() / total_spend * 100).round(1)
        sup["bar_colour"] = sup["cum_pct"].apply(
            lambda p: "#ef4444" if p <= 50 else ("#f59e0b" if p <= 80 else "#22c55e")
        )
        sup_labels  = sup["Vendor"].fillna("Unknown").tolist()
        sup_values  = sup["Spend_MYR"].tolist()
        sup_colours = sup["bar_colour"].tolist()
    else:
        sup_labels = sup_values = sup_colours = []

    chart_data = {
        "donut": {"labels": donut_labels, "values": donut_values, "colours": donut_colours},
        "suppliers": {"labels": sup_labels, "values": sup_values, "colours": sup_colours},
    }

    return render_template(
        "dashboard.html",
        total=total, scope1=scope1, scope2=scope2, scope3=scope3,
        chart_data=json.dumps(chart_data),
        warnings=warnings,
        record_count=len(df),
    )


@app.route("/api/records")
def api_records():
    sid = session.get("ghg_sid")
    if not sid:
        return jsonify([])

    df = load_session(sid)
    if df is None:
        return jsonify([])

    df = df.copy()
    df["Date"]      = df["Date"].astype(str)
    df["Spend_MYR"] = df["Spend_MYR"].fillna(0)
    df["Qty"]       = df["Qty"].fillna(0)
    df["EF_val"]    = df["EF_val"].fillna(0)

    records = df.where(pd.notna(df), None).to_dict(orient="records")
    return jsonify(records)


@app.route("/export")
def export():
    sid = session.get("ghg_sid")
    if not sid:
        return redirect(url_for("index"))

    df = load_session(sid)
    if df is None:
        return redirect(url_for("index"))

    output = io.BytesIO()
    _write_excel(df, output)
    output.seek(0)

    return send_file(
        output,
        download_name=f"AETH_GHG_Report_{REPORTING_YEAR}.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# ── EXCEL WRITER ──────────────────────────────────────────────────────────────
def _write_excel(df, output):
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    HDR_FILL  = PatternFill("solid", fgColor="0F172A")
    HDR_FONT  = Font(color="FFFFFF", bold=True)
    ALT_FILL  = PatternFill("solid", fgColor="F1F5F9")
    BOLD      = Font(bold=True)

    def _style_sheet(ws, data_df):
        cols = list(data_df.columns)
        ws.append(cols)
        for cell in ws[1]:
            cell.fill  = HDR_FILL
            cell.font  = HDR_FONT
            cell.alignment = Alignment(horizontal="center")
        ws.freeze_panes = "A2"

        for i, row in enumerate(data_df.itertuples(index=False), start=2):
            ws.append(list(row))
            if i % 2 == 0:
                for cell in ws[i]:
                    cell.fill = ALT_FILL

        for col_idx, col_name in enumerate(cols, start=1):
            max_len = max(
                len(str(col_name)),
                *(len(str(v)) for v in data_df[col_name].astype(str))
            ) + 2
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len, 50)

    wb = Workbook()
    wb.remove(wb.active)

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")

    scope_totals = (
        df.groupby("Scope")["tCO2e"].sum()
          .reset_index()
          .rename(columns={"Scope": "Scope", "tCO2e": "Total tCO2e"})
    )
    scope_totals.loc[len(scope_totals)] = ["TOTAL", round(df["tCO2e"].sum(), 4)]

    ws_sum.append(["Scope", "Total tCO2e"])
    for cell in ws_sum[1]:
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
    for row in scope_totals.itertuples(index=False):
        ws_sum.append(list(row))

    ws_sum.append([])
    ws_sum.append(["Supplier Spend Ranking (SAP PO)"])
    ws_sum[ws_sum.max_row][0].font = BOLD

    po_df = df[df["Source"] == "SAP PO"]
    if not po_df.empty:
        sup = (
            po_df.groupby("Vendor")["Spend_MYR"]
                 .sum()
                 .sort_values(ascending=False)
                 .reset_index()
        )
        sup.columns = ["Vendor", "Spend_MYR"]
        ws_sum.append(["Vendor", "Spend MYR"])
        for cell in ws_sum[ws_sum.max_row]:
            cell.fill = HDR_FILL
            cell.font = HDR_FONT
        for row in sup.itertuples(index=False):
            ws_sum.append(list(row))

    ws_sum.freeze_panes = "A2"
    ws_sum.column_dimensions["A"].width = 40
    ws_sum.column_dimensions["B"].width = 20

    # ── Scoped sheets ─────────────────────────────────────────────────────────
    sheet_defs = [
        ("Scope 1 — Direct",          "Scope 1",  None),
        ("Scope 2 — Electricity",     "Scope 2",  None),
        ("Scope 3 Cat1 — Purchased",  "Scope 3",  "Scope 3 Cat 1 — Purchased Goods"),
        ("Scope 3 Cat5 — Waste",      "Scope 3",  "Scope 3 Cat 5 — Waste / Water"),
        ("Scope 3 Cat7 — Commuting",  "Scope 3",  "Scope 3 Cat 7 — Employee Commuting"),
        ("Scope 3 Cat9 — Transport",  "Scope 3",  "Scope 3 Cat 9 — Downstream Transport"),
        ("Scope 3 Cat11 — Products",  "Scope 3",  "Scope 3 Cat 11 — Use of Sold Products"),
    ]

    for sheet_name, scope_val, cat_val in sheet_defs:
        mask = df["Scope"] == scope_val
        if cat_val:
            mask &= df["GHG_Category"] == cat_val
        sub = df[mask].copy()
        ws  = wb.create_sheet(sheet_name)
        if sub.empty:
            ws.append(["No data for this category"])
        else:
            _style_sheet(ws, sub)

    # ── Supplier analysis ─────────────────────────────────────────────────────
    ws_sup = wb.create_sheet("Supplier Analysis")
    if not po_df.empty:
        sup_detail = (
            po_df.groupby(["Vendor", "GHG_Category"])
                 .agg(Total_Spend_MYR=("Spend_MYR", "sum"),
                      Total_kgCO2e=("kgCO2e", "sum"),
                      Total_tCO2e=("tCO2e", "sum"))
                 .reset_index()
        )
        _style_sheet(ws_sup, sup_detail)
    else:
        ws_sup.append(["No PO data available"])

    wb.save(output)


if __name__ == "__main__":
    app.run(debug=True, port=5000, use_reloader=False)
